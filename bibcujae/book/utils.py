from openpyxl import Workbook
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination


def writeJsonToExcel(data, output_file):
    # Parse the JSON data
    # data = json.loads(json_data)

    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Write the headers to the worksheet
    headers = list(data[0].keys())
    for i, header in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=header)

    # Write the data to the worksheet
    for row, record in enumerate(data, start=2):
        for col, value in enumerate(record.values(), start=1):
            ws.cell(row=row, column=col, value=value)

    # Save the workbook to the specified output file
    wb.save(output_file)


def validateFilters(sf):
    if sf['category'] == 'anno_pub' and sf['option'] not in ['greater_than', 'less_than', 'range', 'exact']:
        return "Las opciones para la regla 'Año de publicación' deben ser 'antes de', 'después de' o 'entre'."
    if sf['category'] != 'anno_pub' and sf['option'] in ['greater_than', 'less_than', 'range']:
        return "Las opciones 'antes de', 'después de' o 'entre' solo se permiten para la categoría 'Año de publicación.'"


class customPagination(PageNumberPagination):
    page_size_query_param = 'items'
    max_page_size = 100

    def get_paginated_response(self, data):
        return Response({
            'count': self.page.paginator.count,
            'current_page': self.page.number,
            'num_pages': self.page.paginator.num_pages,
            'next': self.get_next_link(),
            'previous': self.get_previous_link(),
            'results': data
        })
